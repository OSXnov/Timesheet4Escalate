import tkinter as tk
from tkinter import messagebox, scrolledtext
from tkcalendar import Calendar
import shutil
import os
import openpyxl
from datetime import datetime


def insert_date_into_excel(date_value, file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        sheet['B15'] = date_value  
        workbook.save(file_path)
        messagebox.showinfo("Success", f"Date {date_value} has been inserted into cell B15.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while inserting the date: {e}")


def open_calendar(file_path):
    calendar_window = tk.Toplevel(root)
    calendar_window.title("Select Date for TimeSheet Period")
    calendar_window.geometry("300x250")

    calendar = Calendar(calendar_window, selectmode='day', date_pattern='mm/dd/yyyy')
    calendar.pack(padx=20, pady=20)

    def select_date():
        selected_date = calendar.get_date()
        try:
            date_value = datetime.strptime(selected_date, '%m/%d/%Y').strftime('%Y-%m-%d')
            insert_date_into_excel(date_value, file_path)
            calendar_window.destroy()
        except ValueError:
            messagebox.showerror("Error", "Invalid date format selected.")

    submit_button = tk.Button(calendar_window, text="SUBMIT", command=select_date, font=("Arial", 10))
    submit_button.pack(pady=10)


def start_renaming():
    def open_name_window(new_file_path):
        name_window = tk.Toplevel(root)
        name_window.title("Set Name for Timesheet")
        name_window.geometry("400x400")

        def submit_name():
            user_name = name_entry.get()
            if user_name:
                try:
                    workbook = openpyxl.load_workbook(new_file_path)
                    sheet = workbook.active
                    sheet['C7'] = user_name
                    workbook.save(new_file_path)
                    messagebox.showinfo("Success", f"Name '{user_name}' has been set in cell C7.")
                    name_window.destroy()
                    open_day_input_window(new_file_path)
                except Exception as e:
                    messagebox.showerror("Error", f"Could not write name to file: {e}")
            else:
                messagebox.showerror("Error", "Please enter a name.")

        set_name_label = tk.Label(name_window, text="Enter your name:", font=("Arial", 10))
        set_name_label.pack(pady=10)

        name_entry = tk.Entry(name_window, width=30)
        name_entry.pack(pady=10)

        submit_name_button = tk.Button(name_window, text="SUBMIT", command=submit_name, font=("Arial", 10))
        submit_name_button.pack(pady=10)

        set_period_button = tk.Button(
            name_window, text="Set TimeSheet Period", font=("Arial", 10),
            command=lambda: open_calendar(new_file_path)
        )
        set_period_button.pack(pady=10)

        close_button = tk.Button(name_window, text="CLOSE", font=("Arial", 10), command=name_window.destroy)
        close_button.pack(pady=10)

    rename_window = tk.Toplevel(root)
    rename_window.title("Rename Timesheet File")
    rename_window.geometry("300x300")

    label = tk.Label(rename_window, text="Enter the new file name:", font=("Arial", 10))
    label.pack(pady=10)

    file_name_entry = tk.Entry(rename_window, width=30)
    file_name_entry.pack(pady=10)

    def submit_file_name():
        new_file_name = file_name_entry.get()
        original_file = 'TimesheetTemplate.xlsx'
        new_file_path = f'TimesheetStorage/{new_file_name}.xlsx'

        if os.path.exists(original_file):
            try:
                shutil.copyfile(original_file, new_file_path)
                messagebox.showinfo("Success", f"File renamed to {new_file_name}.xlsx successfully.")
                rename_window.destroy()
                open_name_window(new_file_path)
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")
        else:
            messagebox.showerror("Error", "Original file not found!")

    submit_button = tk.Button(rename_window, text="SUBMIT", command=submit_file_name, font=("Arial", 10))
    submit_button.pack(pady=10)


def open_day_input_window(file_path):
    day_input_window = tk.Toplevel(root)
    day_input_window.title("Daily Timesheet Entry")
    day_input_window.geometry("500x500")

 
    description_slots = ['C15', 'C18', 'C21', 'C24', 'C27', 'C30', 'C33',
                         'I15', 'I18', 'I21', 'I24', 'I27', 'I30', 'I33']
    hours_slots = ['E15', 'E18', 'E21', 'E24', 'E27', 'E30', 'E33',
                   'J15', 'J18', 'J21', 'J24', 'J27', 'J30', 'J33']

    total_hours = 0
    current_day_index = 0  


    def load_workbook():
        try:
            return openpyxl.load_workbook(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load Excel file: {e}")
            return None

    workbook = load_workbook()
    if not workbook:
        day_input_window.destroy()
        return

    sheet = workbook.active


    hours_list = []

 
    def update_excel(description, hours, day_index):
        try:
            desc_slot = description_slots[day_index]
            hrs_slot = hours_slots[day_index]
            sheet[desc_slot] = description
            sheet[hrs_slot] = hours
            workbook.save(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Could not update Excel file: {e}")


    def toggle_worked():
        if not_worked_var.get():
            description_text.config(state='disabled')
            hours_entry.config(state='disabled')
        else:
            description_text.config(state='normal')
            hours_entry.config(state='normal')

    
    def next_day():
        nonlocal current_day_index, total_hours
        if not_worked_var.get():
            description = ""
            hours = 0
        else:
            description = description_text.get('1.0', tk.END).strip()
            try:
                hours = float(hours_entry.get())
                hours_list.append(hours)
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number for hours.")
                return

        update_excel(description, hours, current_day_index)
        current_day_index += 1

        if current_day_index < len(description_slots):
            load_day()
        else:
            
            total_hours = sum(hours_list)
            messagebox.showinfo("Total Hours", f"Total Hours Worked: {total_hours}")
            day_input_window.destroy()

    
    def submit_data():
        nonlocal total_hours
        if not_worked_var.get():
            description = ""
            hours = 0
        else:
            description = description_text.get('1.0', tk.END).strip()
            try:
                hours = float(hours_entry.get())
                hours_list.append(hours)
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number for hours.")
                return

        update_excel(description, hours, current_day_index)
        total_hours = sum(hours_list)
        messagebox.showinfo("Total Hours", f"Total Hours Worked: {total_hours}")
        day_input_window.destroy()

    
    def load_day():
        day_label.config(text=f"Day {current_day_index + 1} of {len(description_slots)}")
        not_worked_var.set(0)
        toggle_worked()

        description_text.config(state='normal')
        hours_entry.config(state='normal')

        description_text.delete('1.0', tk.END)
        hours_entry.delete(0, tk.END)

    
    day_label = tk.Label(day_input_window, text=f"Day {current_day_index + 1} of {len(description_slots)}",
                         font=("Arial", 12))
    day_label.pack(pady=10)

    worked_frame = tk.Frame(day_input_window)
    worked_frame.pack(pady=5)

    not_worked_var = tk.IntVar()
    worked_checkbox = tk.Checkbutton(worked_frame, text="Didn't work?", variable=not_worked_var,
                                     command=toggle_worked, font=("Arial", 10))
    worked_checkbox.pack()

    description_label = tk.Label(day_input_window, text="Description of the day:", font=("Arial", 10))
    description_label.pack(pady=5)

    description_text = scrolledtext.ScrolledText(day_input_window, wrap=tk.WORD, width=40, height=4)
    description_text.pack(pady=5)

    hours_label = tk.Label(day_input_window, text="Hours worked:", font=("Arial", 10))
    hours_label.pack(pady=5)

    hours_entry = tk.Entry(day_input_window, width=10)
    hours_entry.pack(pady=5)

    next_button = tk.Button(day_input_window, text="NEXT DAY", command=next_day, font=("Arial", 10))
    next_button.pack(pady=10)

    submit_button = tk.Button(day_input_window, text="SUBMIT", command=submit_data, font=("Arial", 10))
    submit_button.pack(pady=10)

    
    load_day()


root = tk.Tk()
root.title("Timesheet Management System")
root.geometry("300x200")


start_button = tk.Button(root, text="Start Timesheet Process", command=start_renaming, font=("Arial", 12))
start_button.pack(pady=50)


root.mainloop()
